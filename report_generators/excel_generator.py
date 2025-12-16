# report_generators/excel_generator.py
import pandas as pd
import logging
import os
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import utils

def _format_excel_sheet(worksheet, has_total_row=False):
    """Áp dụng định dạng (in đậm, kẻ khung, tự động giãn cột) cho một worksheet."""
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for cell in worksheet[1]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    if has_total_row and max_row > 1:
        for cell in worksheet[max_row]:
            cell.font = bold_font

    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def _apply_conditional_formatting(worksheet, highlight_config):
    """Tô màu các dòng có giá trị trong cột chỉ định vượt ngưỡng."""
    if not highlight_config or 'threshold' not in highlight_config or 'column_name' not in highlight_config:
        return

    threshold = highlight_config['threshold']
    column_name = highlight_config['column_name']
    red_fill = PatternFill(start_color="FFFFCDCD", end_color="FFFFCDCD", fill_type="solid")
    
    days_col_idx = -1
    for idx, cell in enumerate(worksheet[1]):
        if cell.value == column_name:
            days_col_idx = idx + 1
            break
    
    if days_col_idx == -1:
        logging.warning(f"Không tìm thấy cột '{column_name}' để áp dụng định dạng có điều kiện.")
        return

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        days_cell = row[days_col_idx - 1]
        try:
            if int(days_cell.value) > threshold:
                for cell in row:
                    cell.fill = red_fill
        except (ValueError, TypeError):
            continue

def generate_excel_report(path, data, columns_map, total_row=None, highlight_config=None):
    """
    Hàm chung để xuất dữ liệu ra file Excel với định dạng chuyên nghiệp.
    """
    if not data and not total_row:
        return {"success": False, "message": "Không có dữ liệu để xuất"}
    try:
        processed_data = [dict(row) for row in data]

        date_columns = ['date_in', 'date_out']
        for row in processed_data:
            for col_name in date_columns:
                if col_name in row:
                    row[col_name] = utils.format_datetime_for_display(row.get(col_name))
        
        df_data = []
        for item in processed_data:
            row_data = {key: item.get(key) for key in columns_map.keys()}
            df_data.append(row_data)

        for i, row in enumerate(df_data, 1):
            row['stt'] = i
        
        if total_row:
            total_row['stt'] = ''
            df_data.append(total_row)

        df = pd.DataFrame(df_data)
        
        final_cols_map = {'stt': 'STT', **columns_map}
        ordered_cols = [key for key in final_cols_map.keys() if key in df.columns]
        df = df[ordered_cols]
        df = df.rename(columns=final_cols_map)
        
        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                
                _format_excel_sheet(worksheet, has_total_row=(total_row is not None))
                
                if highlight_config:
                    _apply_conditional_formatting(worksheet, highlight_config)
        except IOError:
            logging.error(f"Lỗi ghi file Excel tại '{path}'. Có thể file đang được mở.")
            return {"success": False, "message": f"Không thể ghi file.\nVui lòng đảm bảo file '{os.path.basename(path)}' không đang được mở."}

        return {"success": True, "message": "Xuất file thành công."}
    except Exception as e:
        logging.exception("Lỗi không xác định khi xuất ra Excel")
        return {"success": False, "message": str(e)}